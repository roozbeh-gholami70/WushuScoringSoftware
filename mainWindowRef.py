import time
import inspect
from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QSplashScreen 
from PyQt5 import QtCore, QtGui , QtSerialPort
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtMultimedia import QMediaPlayer
from PyQt5.QtCore import Qt
from PyQt5.uic import loadUiType
import winsound
import resources_rc


import string
from score import  Main

import openpyxl 
from openpyxl.styles import PatternFill
import pandas as pd

import sys
import os
import copy


def loadUiClass(path):
    stream = QFile(path)
    stream.open(QFile.ReadOnly)
    try:
        return loadUiType(stream)[0]
    finally:
        stream.close()

# pyrcc5 resources.qrc -o resources_rc.py
# qt keys https://web.mit.edu/~firebird/arch/sun4x_59/doc/html/qt.html

t = 0

redLcdList = ["lcdNumber", "lcdNumber_3", "lcdNumber_7","lcdNumber_9", "lcdNumber_13"]
blueLcdList = ["lcdNumber_2", "lcdNumber_4", "lcdNumber_8", "lcdNumber_10", "lcdNumber_14"]
colorDict = {"red":0, "blue":1, "white":2}

tableFirstColList = ["داور1","داور2","داور3","داور4","داور5","اخطار","ناکدان", "ناک اوت","تذکر","خروج"]
tableHeaderList = ["نام","راند","داور1","داور2","داور3","داور4","داور5","اخطار","ناکدان", "ناک اوت","تذکر","خروج"]
alphabetList = list(string.ascii_uppercase)


class TimerDialog(QDialog, loadUiClass(':/ui_files/timerDialog.ui')):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)

class ControllerDialog(QDialog, loadUiClass(':/ui_files/controllerDialog.ui')):
    def __init__(self, serialConn, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.controllerIdxList = []
        self.controllerMacList = []
        for idx in range(1,6):
            myObjCbx = eval("self.controlllerCbx_"+str(idx))
            myObjCbx.clear()
            myObjCbx.addItem("__")
            self.controllerIdxList.append(None)
            myObjCbx.setCurrentText("__")
            items  = list(range(1, 6))
            myObjCbx.addItems(map(str, items))
        self.controlllerCbx_1.activated.connect(lambda:self.controlllerSelection(self.controlllerCbx_1))
        self.controlllerCbx_2.activated.connect(lambda:self.controlllerSelection(self.controlllerCbx_2))
        self.controlllerCbx_3.activated.connect(lambda:self.controlllerSelection(self.controlllerCbx_3))
        self.controlllerCbx_4.activated.connect(lambda:self.controlllerSelection(self.controlllerCbx_4))
        self.controlllerCbx_5.activated.connect(lambda:self.controlllerSelection(self.controlllerCbx_5))
        

        
    def updateComboBoxes(self):
        for idx in range(1,6):
            if (self.controllerIdxList[idx-1]==None):
                myObjCbx = eval("self.controlllerCbx_"+str(idx))
                myObjCbx.clear()
                myObjCbx.addItem("__")
                myObjCbx.setCurrentText("__")
                items = []
                for tmp in range(1, 6):
                    if tmp not in self.controllerIdxList:
                        items.append(tmp)
                myObjCbx.addItems(map(str, items))
            else:
                myObjCbx = eval("self.controlllerCbx_"+str(idx))
                myObjCbx.clear()
                items = []
                items.append("__")
                for tmp in range(1, 6):
                    if tmp not in self.controllerIdxList:
                        items.append(tmp)
                items.append(self.controllerIdxList[idx-1])
                myObjCbx.addItems(map(str, items))
                myObjCbx.setCurrentText(str(self.controllerIdxList[idx-1]))

    def controlllerSelection(self, myCbx):
        myCbxIdx = int(myCbx.objectName()[-1])-1
        if (myCbx.currentText().isdigit()):
            self.controllerIdxList[myCbxIdx] = int(myCbx.currentText())
            self.updateComboBoxes()
        else:
            self.controllerIdxList[myCbxIdx] = None
            self.updateComboBoxes()

        
    
    def getControllerVal(self, value):
        if value[2] not in self.controllerMacList:
            self.controllerMacList.append(value[2][1:16])
            
        idx = int(self.controllerMacList.index(value[2][1:16])) + 1
        
        myObjLbl = eval("self.red_"+str(idx))
        myObjLbl.setText(value[0])
        myObjLbl = eval("self.blue_"+str(idx))
        myObjLbl.setText(value[1])


class MainWindow (QMainWindow, loadUiClass(':/ui_files/MainWindowReferee.ui')):
    def __init__(self):
        super( MainWindow, self).__init__()
        id = QFontDatabase.addApplicationFont(":/fonts/BYekan.ttf")
        if id < 0: print("Error")  
        families = QFontDatabase.applicationFontFamilies(id)
        QApplication.instance().setFont(QFont(families[0]))
        self.setupUi(self)
        self.player = QMediaPlayer()
        self.show()
        porttnavn = pyqtSignal(str)
        
        self.dfResult = pd.DataFrame()
        self.pushButton_5.clicked.connect(lambda:self.up1r())
        self.pushButton_9.clicked.connect(lambda:self.up2r())
        self.pushButton_22.clicked.connect(lambda:self.up3r())
        self.pushButton_28.clicked.connect(lambda:self.up4r())
        self.pushButton_36.clicked.connect(lambda:self.up5r())

        self.pushButton_7.clicked.connect(lambda:self.up1b())
        self.pushButton_14.clicked.connect(lambda:self.up2b())
        self.pushButton_23.clicked.connect(lambda:self.up3b())
        self.pushButton_26.clicked.connect(lambda:self.up4b())
        self.pushButton_34.clicked.connect(lambda:self.up5b())

        ##########################################################
        self.pushButton_6.clicked.connect(lambda:self.dn1r())
        self.pushButton_12.clicked.connect(lambda:self.dn2r())
        self.pushButton_24.clicked.connect(lambda:self.dn3r())
        self.pushButton_27.clicked.connect(lambda:self.dn4r())
        self.pushButton_35.clicked.connect(lambda:self.dn5r())

        self.pushButton_8.clicked.connect(lambda:self.dn1b())
        self.pushButton_13.clicked.connect(lambda:self.dn2b())
        self.pushButton_15.clicked.connect(lambda:self.dn3b())
        self.pushButton_25.clicked.connect(lambda:self.dn4b())
        self.pushButton_33.clicked.connect(lambda:self.dn5b())
        
        
        self.timeAction.triggered.connect(self.openTimerDialog)
        self.controllerAction.triggered.connect(self.openControllerDialog)

        self.pushButton_3.setEnabled(False)
        self.endBtn.setEnabled(False)
        self.sendBtn.setEnabled(False)
        self.connectBtn.setEnabled(False)
        self.timerStartBtn.clicked.connect(lambda:self.start1())
        self.pushButton_3.clicked.connect(lambda:self.puse())
        self.endBtn.clicked.connect(lambda:self.resetTime())
        self.endBtn.clicked.connect(lambda:self.receive())
        self.connectBtn.clicked.connect(lambda:self.on_toggled())
        # self.connectBtn.clicked.connect(lambda:self.on_toggled())
        self.sendBtn.clicked.connect(lambda:self.sendToArdo())
        # self.endBtn.clicked.connect(lambda:self.choose_music())

        self.freeBtn.clicked.connect(lambda:self.sendToArdo(packet="944444"))
        self.lockBtn.clicked.connect(lambda:self.sendToArdo(packet="999999"))

        self.uploadBtn.clicked.connect(lambda:self.uploadData())
        



        self.weightTxt.textChanged.connect(lambda:self.setWeight())
        
        self.gameResetBtn.clicked.connect(self.resetGame)
        self.gameNumCbx.currentIndexChanged.connect(self.setGameNum)
        self.portsCbx.activated.connect(lambda:self.setPortName())
        

        self.scoreBoardBtn.clicked.connect(lambda:self.openSecondWindow())
        self.scoreBoardAction.triggered.connect(lambda:self.openSecondWindow())

        
        self.playerNameAction.triggered.connect(lambda:self.takePlayerNames())


        
        self.gameNumAction.triggered.connect(lambda:self.getMatchSettings())

        
        self.saveCsvAction.triggered.connect(lambda:self.saveToCsv())

        self.round1Rbtn.clicked.connect(lambda:self.radioBtnState(self.round1Rbtn))
        self.round2Rbtn.clicked.connect(lambda:self.radioBtnState(self.round2Rbtn))
        self.round3Rbtn.clicked.connect(lambda:self.radioBtnState(self.round3Rbtn))

        
        self.uploadFileAction.triggered.connect(lambda: self.loadCsvFile())

        self.gameNumCbx.setStyleSheet("QComboBox QAbstractItemView { \
                                        selection-color: red; \
                                        selection-background-color: rgb(244, 244, 244);}")
        
        
        self.portsCbx.addItem("__")
        self.portsCbx.setCurrentText("__")
        self.portsList = ["__"]
        self._getserial_ports()
        
        self.controlDialog = None

        self.isMatchData = False
        self.loadingCsv = False
        self.round1Rbtn.setChecked(True)
        self.mainSecond = 0
        self.mainMintues = 0
        self.count = 0
        self.minutes = 0
        self.secondarySecond = 0
        self.secondaryMinute = 0
        self.restSecond = 0
        self.restMinute = 0
        self.roundNum = 1
        self.numOfMatches = 0
        self.start = False
        self.mainTime = False
        self.restTime = False
        self.lockRef = True
        self.redPoints = 0
        self.bluePoints = 0
        self.redWinRand = [0, 0 , 0]
        self.blueWinRand = [0, 0 , 0]
        self.gameNum = ""
        self.redPlayer = ""
        self.bluePlayer = ""
        self.playerWeight = ""
        self.portName = "__"
        self.serial = QtSerialPort.QSerialPort(self.portName,
                                                baudRate=QtSerialPort.QSerialPort.Baud115200, 
                                                readyRead=self.receive)
        self.refdic = {"5c:cf:7f:c9:1b:":1 , "48:3f:da:0f:92:":2, "8c:aa:b5:cf:bc:" :3,  "30:83:98:a6:5c:":4 , "5c:cf:7f:96:30:":5}

        self.refColor = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
        
        self.redPlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
        self.bluePlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}

        self.redPlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
        self.bluePlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
        # my part start
        self.setSpinBoxReadOnly()
        timer = QTimer(self)
        # adding action to timer
        timer.timeout.connect(self.showTime)

        # update the timer every tenth second
        timer.start(1000)

        # my part end
    
        self.minet = 0
        self.t =0
    
    def openControllerDialog(self):
        if self.serial.isOpen():
            self.controlDialog = ControllerDialog(self.serial, self)
            if (self.controlDialog.exec_()==QDialog.Accepted):
                if (None not in self.controlDialog.controllerIdxList):
                    self.sendToArdo(packet="922222")
                    self.refdic = dict(map(lambda i,j : (i,j) , self.controlDialog.controllerMacList,self.controlDialog.controllerIdxList))
                    QMessageBox.about(self, "اعلان", "تنظیمات با موفقیت انجام شد.")
                    print(self.refdic)
                else:
                    QMessageBox.about(self, "اخطار", "لطفا تنظیمات را برای تمامی دسته ها انجام دهید.")
            else:
                print("Canceled")
        else:
            QMessageBox.about(self, "اخطار", "لطفا اتصال سریال را برقرار کنید.")


    def openTimerDialog(self):
        dialog = TimerDialog(self)
        if (dialog.exec_()==QDialog.Accepted):
            self.start = False
            second = int(dialog.gameTime.time().second())
            minute = int(dialog.gameTime.time().minute())
            self.mainMintues = minute
            self.mainSecond = second

            self.label_2.setText("{:02d}".format(minute))
            self.label_8.setText("{:02d}".format( second))
            self.pushButton_3.setEnabled(False)
            self.pushButton_3.setText("وقفه")
            self.timerStartBtn.setEnabled(True)
            self.second = second
            self.secs = second * 10
            self.count = second
            self.minutes = minute
            self.mainTime = True
            restSecond = int(dialog.restTime.time().second())
            restMinute = int(dialog.restTime.time().minute())
            if ((restSecond != 0) | (restMinute != 0)):
                self.restTime = True
                self.label_16.setText("{:02d}".format(restSecond))
                self.label_14.setText("{:02d}".format(restMinute ))
                self.restSecond = restSecond
                self.restMinute = restMinute
                self.secondarySecond = restSecond
                self.secondaryMinute = restMinute
            else:
                self.restTime = False



    def setSpinBoxReadOnly(self):
        for roundNum in range(1,4):
            for idx in range(5):
                myObj = eval("self.red"+str(roundNum)+"_"+str(idx+1) + "spinBox")
                myObj.lineEdit().setReadOnly(True)

                myObj = eval("self.blue"+str(roundNum)+"_"+str(idx+1) + "spinBox")
                myObj.lineEdit().setReadOnly(True)
        self.makeDisableEnable()

                
        
    def keyPressEvent(self, e):

        modifiers = QtWidgets.QApplication.keyboardModifiers()
      
        if e.modifiers() & Qt.ControlModifier:
            if e.key() == Qt.Key_7:
                myObj = eval("self.red"+str(self.roundNum)+"_1spinBox")
                if (int(myObj.value())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp)         

            if e.key() == Qt.Key_4:
                myObj = eval("self.red"+str(self.roundNum)+"_2spinBox")
                if (int(myObj.text())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp)    
            
            if e.key() == Qt.Key_1:
                myObj = eval("self.red"+str(self.roundNum)+"_3spinBox")
                if (int(myObj.text())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp)    
            if e.key() == Qt.Key_0:
                myObj = eval("self.red"+str(self.roundNum)+"_4spinBox")
                if (int(myObj.text())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp)                           

            if e.key() == Qt.Key_9:
                myObj = eval("self.blue"+str(self.roundNum)+"_1spinBox")
                if (int(myObj.text())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp)    

            if e.key() == Qt.Key_6:
                myObj = eval("self.blue"+str(self.roundNum)+"_2spinBox")
                if (int(myObj.text())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp)
            
            if e.key() == Qt.Key_3:
                myObj = eval("self.blue"+str(self.roundNum)+"_3spinBox")
                if (int(myObj.text())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp) 
            if e.key() == Qt.Key_Period:
                myObj = eval("self.blue"+str(self.roundNum)+"_4spinBox")
                if (int(myObj.text())>0):
                    tmp = int(myObj.value()) - 1
                    myObj.setValue(tmp) 
        else:
            if e.key() == Qt.Key_F1:
                if (self.timerStartBtn.isEnabled()):
                    self.start1()
                elif ((self.pushButton_3.isEnabled()) & (self.count != 0)):
                    self.puse()

            if e.key() == Qt.Key_7:
                myObj = eval("self.red"+str(self.roundNum)+"_1spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp)        

            if e.key() == Qt.Key_4:
                myObj = eval("self.red"+str(self.roundNum)+"_2spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp) 
            
            if e.key() == Qt.Key_1:
                myObj = eval("self.red"+str(self.roundNum)+"_3spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp) 
                
            if e.key() == Qt.Key_0:
                myObj = eval("self.red"+str(self.roundNum)+"_4spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp) 

            if e.key() == Qt.Key_9:
                myObj = eval("self.blue"+str(self.roundNum)+"_1spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp) 

            if e.key() == Qt.Key_6:
                myObj = eval("self.blue"+str(self.roundNum)+"_2spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp) 
            
            if e.key() == Qt.Key_3:
                myObj = eval("self.blue"+str(self.roundNum)+"_3spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp) 
            
            if e.key() == Qt.Key_Period:
                myObj = eval("self.blue"+str(self.roundNum)+"_4spinBox")
                tmp = int(myObj.value()) + 1
                myObj.setValue(tmp)  

        
    def takePlayerNames(self):
        self.redPlayerData = {}
        self.bluePlayerData = {}

        if (self.numOfMatches == 0):
            QMessageBox.about(self, "اخطار", "لطفا تعداد مسابقات را مشخص کنید.")
        else:
            redPlayerName, done1 = QtWidgets.QInputDialog.getText(
                self, 'مشخصات مبارز قرمز', 'نام مبارز قرمز:')
    
            bluePlayerName, done2 = QtWidgets.QInputDialog.getText(
            self, 'مشخصات مبارز آبی', 'نام مبارز آبی:') 


            if done1 and done2 :
                msgBoxTopic = "مشخصات مبارزین"
                msgBoxQuestion = "مبارز قرمز: " + redPlayerName+ "\n" +"مبارز آبی: " + bluePlayerName + "\n"
                buttonReply = QMessageBox.question(self,msgBoxTopic ,msgBoxQuestion , QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
                if buttonReply == QMessageBox.Yes:
                    self.redPlayer = redPlayerName
                    self.bluePlayer = bluePlayerName
                    self.redPlayerData["name"] = redPlayerName
                    self.bluePlayerData["name"] = bluePlayerName
                    self.redPlayerData["weight"] = str(self.playerWeight)
                    self.bluePlayerData["weight"] = str(self.playerWeight)
                    self.redPlayerData["matchNum"] = str(self.gameNum)
                    self.bluePlayerData["matchNum"] = str(self.gameNum)
                    self.setPlayerName()
                else:
                    self.takePlayerNames()
    
    def getMatchSettings(self):
        numOfMatches, done1 = QtWidgets.QInputDialog.getText(
            self, '  تعداد مسابقات', 'تعداد مسابقه: ')
        if done1:
            msgBoxTopic = "تعداد مسابقات"
            msgBoxQuestion = " تعداد مسابقه: "+ numOfMatches + "\n"
            buttonReply = QMessageBox.question(self,msgBoxTopic ,msgBoxQuestion , QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if buttonReply == QMessageBox.Yes:
                self.numOfMatches = int(numOfMatches)
                self.redPlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                self.bluePlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}

                self.redPlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                self.bluePlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                self.myDataTable = list()
                for idx in range(int(numOfMatches)):
                    tmp = {}
                    tmp["redRefPoint"] = copy.deepcopy(self.redPlayerRefsPoint)
                    tmp["redTablePoint"] = copy.deepcopy(self.redPlayerTablePoint)
                    tmp["blueRefPoint"] = copy.deepcopy(self.bluePlayerRefsPoint)
                    tmp["blueTablePoint"] = copy.deepcopy(self.bluePlayerTablePoint)
                    self.myDataTable.append(copy.deepcopy(tmp))
                
                self.gameNumCbx.clear()
                items  = list(range(1, int(numOfMatches)+1))
                self.gameNumCbx.addItems(map(str, items))
                self.gameNum = self.gameNumCbx.currentText()
                self.isMatchData = True
                self.resetRefLcd()
                self.setMatchResult()
            else:
                self.getMatchSettings()


    def setTable(self):
        for i in range(self.gridTable.count()):
            item = self.gridTable.itemAt(i).widget()
            if isinstance(item,QSpinBox):
                item.setValue(0)

    def makeDisableEnable(self):
        roundsSet = set(["1","2","3"])-set(str(self.roundNum))
        for idx in range(5):
            for randIdx in roundsSet:
                myObj = eval("self.red"+str(randIdx)+"_"+str(idx+1) + "spinBox")
                myObj.setEnabled(False)

                myObj = eval("self.blue"+str(randIdx)+"_"+str(idx+1) + "spinBox")
                myObj.setEnabled(False)
            myObj = eval("self.red"+str(self.roundNum)+"_"+str(idx+1) + "spinBox")
            myObj.setEnabled(True)

            myObj = eval("self.blue"+str(self.roundNum)+"_"+str(idx+1) + "spinBox")
            myObj.setEnabled(True)

    def radioBtnState(self,myBtn):
        self.storeTableData()
        self.calcPoints()
        if myBtn.isChecked() == True:
            self.roundNum = int(str(myBtn.objectName())[5])
            self.makeDisableEnable()
            if ("Main" in dir(self)):
                self.Main.roundNumLbl.setText(str(self.roundNum))
            self.setRefLcd()


    def setRefLcd(self):
        for idx, item in enumerate(redLcdList):
            myObjRed = eval("self."+item)
            myObjBlue = eval("self."+blueLcdList[idx])
            myObjRed.display(self.redPlayerRefsPoint[str(self.roundNum)][idx])
            myObjBlue.display(self.bluePlayerRefsPoint[str(self.roundNum)][idx])


    def setTableVal(self):
        for idx in range(5):
            for round in range(1,4):
                myObj = eval("self.red"+str(round)+"_"+str(idx+1) + "spinBox")
                myObj.lineEdit().setReadOnly(False)
                myObj.setValue(int(self.redPlayerTablePoint[str(round)][idx]))
                myObj.lineEdit().setReadOnly(True)

                myObj = eval("self.blue"+str(round)+"_"+str(idx+1) + "spinBox")
                myObj.lineEdit().setReadOnly(False)
                myObj.setValue(int(self.bluePlayerTablePoint[str(round)][idx]))
                myObj.lineEdit().setReadOnly(True)

    def setPlayerName(self):
        isExcelUploaded = False
        if ("Main" in dir(self)):
            try:
                self.matchData
                self.redPlayerData = {}
                self.bluePlayerData = {}
            except:
                self.Main.redPlayerNameTxt.setPlainText(self.redPlayer)
                self.Main.bluePlayerNameTxt.setPlainText(self.bluePlayer)
                pass
            else:
                isExcelUploaded = True
                self.redPlayerData["name"] = self.matchData[str(self.gameNum)]["red"]
                self.bluePlayerData["name"] = self.matchData[str(self.gameNum)]["blue"]
                self.redPlayerData["weight"] = str(self.playerWeight)
                self.bluePlayerData["weight"] = str(self.playerWeight)
                self.redPlayerData["matchNum"] = str(self.gameNum)
                self.bluePlayerData["matchNum"] = str(self.gameNum)

                self.Main.redPlayerNameTxt.setPlainText(self.matchData[str(self.gameNum)]["red"])
                self.Main.bluePlayerNameTxt.setPlainText(self.matchData[str(self.gameNum)]["blue"])
                self.weightTxt.setPlainText(self.matchData[str(self.gameNum)]["weight"])
                
        
            
    
    def setWeight(self):
        self.playerWeight = self.weightTxt.toPlainText()
        try:
            self.redPlayerData["weight"] = str(self.playerWeight)
            self.bluePlayerData["weight"] = str(self.playerWeight)
        except:
            pass
        if ("Main" in dir(self)):
            self.Main.weightLbl.setText(str(self.playerWeight))
           
    
    def setGameNum(self,index):
        preGameNum = copy.deepcopy(self.gameNum)
        if (self.loadingCsv == False):
            changeGame = QtWidgets.QMessageBox.question(self,
                                            "تغییر مسابقه",
                                            "آیا برای تغییر مسابقه اطمینان دارید؟",
                                            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        else:
            changeGame = QtWidgets.QMessageBox.Yes
            self.loadingCsv = False
        if changeGame == QtWidgets.QMessageBox.Yes:
            self.redWinRand = [0, 0 , 0]
            self.blueWinRand = [0, 0 , 0]
            self.calcPoints()
            self.storeTableData()
            new_index = self.gameNumCbx.currentIndex() 
            self.gameNum = self.gameNumCbx.currentText()
            try:
                self.myDataTable[int(preGameNum)-1]["redRefPoint"] = copy.deepcopy(self.redPlayerRefsPoint)
                self.myDataTable[int(preGameNum)-1]["redTablePoint"] = copy.deepcopy(self.redPlayerTablePoint)
                self.myDataTable[int(preGameNum)-1]["blueRefPoint"] = copy.deepcopy(self.bluePlayerRefsPoint)
                self.myDataTable[int(preGameNum)-1]["blueTablePoint"] = copy.deepcopy(self.bluePlayerTablePoint)

                self.redPlayerRefsPoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["redRefPoint"])
                self.redPlayerTablePoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["redTablePoint"])
                self.bluePlayerRefsPoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["blueRefPoint"])
                self.bluePlayerTablePoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["blueTablePoint"])
                self.setRefLcd()
                self.setTableVal()
            except Exception as ex:
                print("line {}: {}".format(str(inspect.currentframe().f_lineno),ex))
                pass
            if ("Main" in dir(self)):
                self.Main.gameNumTxt.setHtml("<p align='center'>" + str(self.gameNum) + "</p>")
                self.setPlayerName()
                for tmpRoundNum in range(1,4):
                    myObj = eval("self.Main.stackedWidgetBlue"+str(tmpRoundNum))
                    myObj.setCurrentIndex(0)
                    myObj = eval("self.Main.stackedWidgetRed"+str(tmpRoundNum))
                    myObj.setCurrentIndex(0)
                self.Main.winnerStackedWidget.setCurrentIndex(1)
                self.winnerStackedWidget.setCurrentIndex(1)
        else:
            self.gameNumCbx.blockSignals(True)  # Block signals temporarily to prevent recursive calls
            self.gameNumCbx.setCurrentIndex(index)  # Set the index back to the previous value
            self.gameNumCbx.setCurrentText(preGameNum)
            self.gameNumCbx.blockSignals(False)  
            
    def setMatchResult(self):
        preGameNum = copy.deepcopy(self.gameNum)            
        self.calcPoints()
        self.storeTableData()
        self.gameNum = self.gameNumCbx.currentText()
        try:
            self.myDataTable[int(preGameNum)-1]["redRefPoint"] = copy.deepcopy(self.redPlayerRefsPoint)
            self.myDataTable[int(preGameNum)-1]["redTablePoint"] = copy.deepcopy(self.redPlayerTablePoint)
            self.myDataTable[int(preGameNum)-1]["blueRefPoint"] = copy.deepcopy(self.bluePlayerRefsPoint)
            self.myDataTable[int(preGameNum)-1]["blueTablePoint"] = copy.deepcopy(self.bluePlayerTablePoint)

            self.redPlayerRefsPoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["redRefPoint"])
            self.redPlayerTablePoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["redTablePoint"])
            self.bluePlayerRefsPoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["blueRefPoint"])
            self.bluePlayerTablePoint = copy.deepcopy(self.myDataTable[int(self.gameNum)-1]["blueTablePoint"])
            self.setRefLcd()
            self.setTableVal()
        except Exception as ex:
            print("line {}: {}".format(str(inspect.currentframe().f_lineno),ex))
            pass
        if ("Main" in dir(self)):
            self.Main.gameNumTxt.setHtml("<p align='center'>" + str(self.gameNum) + "</p>")
            self.setPlayerName()
            for tmpRoundNum in range(1,4):
                myObj = eval("self.Main.stackedWidgetBlue"+str(tmpRoundNum))
                myObj.setCurrentIndex(0)
                myObj = eval("self.Main.stackedWidgetRed"+str(tmpRoundNum))
                myObj.setCurrentIndex(0)
            self.Main.winnerStackedWidget.setCurrentIndex(1)
            self.winnerStackedWidget.setCurrentIndex(1)
    

    def saveToCsv(self):
        if (self.gameNum!=""):
            try:
                excelFile = "match_"+str(self.gameNum)+".xlsx"
                dir_path = os.path.dirname(os.path.realpath(__file__))
                if os.path.exists(dir_path+"\\" + excelFile):
                    self.myExcel = openpyxl.load_workbook("match_"+str(self.gameNum)+".xlsx")
                    self.workSheet =  self.myExcel.active
                else:
                    self.myExcel = openpyxl.Workbook()
                    self.workSheet = self.myExcel.active
                    self.writeExcelFileHeader()
                
                self.writeExcelFile()
                self.myExcel.save(excelFile)


                self.setTable()
                self.redPlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                self.bluePlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}

                self.redPlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                self.bluePlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                index = self.gameNumCbx.currentIndex()
                self.gameNumCbx.model().item(index).setEnabled(False)
                self.gameNumCbx.setCurrentIndex(index+1)
                self.setMatchResult()
                self.redWinRand = [0, 0 , 0]
                self.blueWinRand = [0, 0 , 0]
                print(self.redPlayerData)
                print(self.bluePlayerData)
            except Exception as e:
                if (("redPlayerData" in str(e)) | ("bluePlayerData" in str(e))):
                    QMessageBox.about(self, "خطا", "لطفا مشصخصات مبارزین را وارد کنید.")
        else:
            QMessageBox.about(self, "اخطار", "لطفا تعداد مسابقات را مشخص کنید.")
        
    def writeExcelFileHeader(self):
        red_cell = PatternFill(patternType='solid', fgColor='FB2D01')
        blue_cell = PatternFill(patternType='solid', fgColor='0163FB')
        self.workSheet['B1'] = str(self.redPlayerData["name"])
        self.workSheet['E1'] = str(self.bluePlayerData["name"])

        # self.workSheet['B1'] = str(self.matchData[str(self.gameNum)]["red"])
        # self.workSheet['E1'] = str(self.matchData[str(self.gameNum)]["blue"])
        try:
            for idx,item in enumerate(alphabetList[1:7]):
                self.workSheet[item+"2"] = "راند" + str((idx%3)+1)
                if (idx < 3):
                    self.workSheet[item +'1'].fill = red_cell
                    self.workSheet[item+str(2)].fill = red_cell
                else:
                    self.workSheet[item +'1'].fill = blue_cell
                    self.workSheet[item+str(2)].fill = blue_cell
            for idx,item in enumerate(tableFirstColList):
                self.workSheet["A" + str(idx+3)] = item
            
            self.workSheet.merge_cells('B1:D1')
            self.workSheet.merge_cells('E1:G1')
            
        except Exception as e:
            print(str(e))
            pass
    
    def writeExcelFile(self):
        red_cell = PatternFill(patternType='solid', fgColor='FB2D01')
        blue_cell = PatternFill(patternType='solid', fgColor='0163FB')
        try:
            for randIdx in range(3):
                for idx in range(3, len(tableFirstColList)+3):
                    if (idx < 8):
                        self.workSheet[alphabetList[randIdx+1]+str(idx)] = self.redPlayerRefsPoint[str(randIdx+1)][idx-3]
                        self.workSheet[alphabetList[randIdx+4]+str(idx)] = self.bluePlayerRefsPoint[str(randIdx+1)][idx-3]
                    else:
                        self.workSheet[alphabetList[randIdx+1]+str(idx)] = self.redPlayerTablePoint[str(randIdx+1)][idx-8]
                        self.workSheet[alphabetList[randIdx+4]+str(idx)] = self.bluePlayerTablePoint[str(randIdx+1)][idx-8] 
                    self.workSheet[alphabetList[randIdx+1]+str(idx)].fill = red_cell
                    self.workSheet[alphabetList[randIdx+4]+str(idx)].fill = blue_cell
        except Exception as e:
            print(str(e))
            pass

    def update_gui1(self):
        
        self.label_14.setText("{:02d}".format(self.mints))
        self.label_16.setText("{:02d}".format( self.scnd))
        if ("Main" in dir(self)):
            self.Main.label_3.setText("{:02d}".format(self.mints))
            self.Main.label_5.setText("{:02d}".format( self.scnd))
    
    def updateTimer(self):
        
        self.label_2.setText("{:02d}".format(self.mints))
        self.label_8.setText("{:02d}".format( self.scnd))
        if ("Main" in dir(self)):
            self.Main.label_3.setText("{:02d}".format( self.mints))
            self.Main.label_5.setText("{:02d}".format(self.scnd))


    
    def setRestTime(self):
        second = self.spinBox_6.value()
        minute = self.spinBox_5.value()
        if ((second != 0) | (minute != 0)):
            self.restTime = True
            self.label_16.setText("{:02d}".format(second))
            self.label_14.setText("{:02d}".format(minute ))
        else:
            self.restTime = False
    
    def showTime(self):
        self.setLcdColor()
        self.calcPoints()
        self.storeTableData()
        # checking if flag is true
        if self.start:
            # incrementing the counter
            self.count -= 1
            # timer is completed
            if self.count == -1:
                if (self.minutes > 0):
                    self.minutes -= 1
                    self.count = 59
                else:
                # making flag false
                    self.start = False
                    # setting text to the label
                    self.mints = 0
                    self.scnd = 0
                    self.pushButton_3.setEnabled(False)
                    self.timerStartBtn.setEnabled(True)
                    self.mainTime = False
                    if (self.lockRef):
                        self.lockRef = False
                        QTimer.singleShot(5000,lambda: self.sendToArdo(packet="999999"))
                    if (self.restTime):
                        self.restTimeFun()
                    else:
                        self.setZeroTime()
                        self.setTimerDefault()
                        
            
            # self.count -= 1
            

        if self.start:
            # getting text from count
            self.scnd = int(self.count)#/10)
            self.mints = self.minutes
            if (self.mainTime==True):
                if (self.minutes == 0):
                    if ((self.count < 11) & (self.count > 0)):
                        #if (self.count == 0):
                        self.secondsound()
                    elif (self.count == 0):
                        self.endtime()
                self.updateTimer()
            elif (self.resetTime):
                if (self.minutes == 0):
                    if ((self.count < 16) & (self.count > 0)):
                        #if (self.count == 0):
                        self.secondsound()
                    elif (self.count == 0):
                        self.endtime()
                self.update_gui1()

    def restTimeFun(self):
        second = self.restSecond
        minute = self.restMinute
        if ((second != 0) | (minute != 0)):
            self.mainTime = False
            self.restTime = True
            self.secs = second * 10
            self.count = (second )#* 10) + 9
            self.minutes = minute

            self.start = True
            self.timerStartBtn.setEnabled(False)
            self.pushButton_3.setEnabled(True)
            self.pushButton_3.setText("وقفه")
            self.restSecond = 0
            self.restMinute = 0
        else:
            self.setZeroTime()
            self.setTimerDefault()
    
    def resetTime(self):
        reply = QMessageBox.question(self, 'پایان راند', 'آیا برای اتمام راند اطمینان دارید؟',
				QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            if (self.start):
                self.start = False
                self.count = 0
                if (~self.restTime):
                    self.restTimeFun()
                else:
                    self.setZeroTime()
                    self.setTimerDefault()
            else:
                self.count = 0
                if (~self.restTime):
                    self.restTimeFun()
                else:
                    self.setZeroTime()
                    self.setTimerDefault()

    def setZeroTime(self):
        self.restTime = False
        self.lockRef = True
        self.mainTime = False
        self.timerStartBtn.setEnabled(True)
        self.endBtn.setEnabled(False)
        self.pushButton_3.setEnabled(False)
        self.pushButton_3.setText("وقفه")
        self.count = 0
        self.minutes = 0
        self.mints = 0
        self.scnd = 0
        self.update_gui1()
        self.updateTimer()

    def setTimerDefault(self):
        self.restTime = False
        self.lockRef = True
        self.mainTime = False
        self.timerStartBtn.setEnabled(True)
        self.endBtn.setEnabled(False)
        self.pushButton_3.setEnabled(False)
        self.pushButton_3.setText("وقفه")
        self.count = self.mainSecond
        self.minutes = self.mainMintues
        self.mainTime = True
        self.restSecond = self.secondarySecond
        self.restMinute = self.secondaryMinute
        if ((self.restMinute!=0) | (self.restSecond!=0)):
            self.mints = self.restMinute
            self.scnd = self.restSecond
            self.update_gui1()
            self.restTime = True
        self.mints = self.minutes
        self.scnd = self.count
        self.updateTimer()

    def start1(self):
        self.start = True
        self.redPoints = 0
        self.bluePoints = 0
        if ((self.minutes==0) & (self.count==0)):
                self.start = False
                QMessageBox.about(self, "اخطار", "لطفا زمان خود را تعیین کنید.")
        elif ((self.minutes != 0) & (self.count==0)):
            self.count = 60
            self.minutes = self.minutes-1

        # if self.count == 0:
        #     self.start = False
        #     QMessageBox.about(self, "اخطار", "لطفا زمان خود را تعیین کنید.")
        if ("Main" not in dir(self)):
            self.start = False
            QMessageBox.about(self, "اخطار", "لطفا اسکوربورد را باز کنید.")
        else:
            self.pushButton_3.setEnabled(True)
            self.timerStartBtn.setEnabled(False)
            self.endBtn.setEnabled(True)
        
        # self.timer_start()

    def puse(self):
        if self.start == True:
            self.start = False
            self.pushButton_3.setText("ادامه")
        else:
            self.start = True
            self.pushButton_3.setText("وقفه")




    def secondsound(self):       
        full_file_path = os.path.join(os.getcwd(), 'time.wav')
        winsound.PlaySound(full_file_path, winsound.SND_ASYNC)        
    def endtime(self):
        full_file_path = os.path.join(os.getcwd(), 'end.wav')
        winsound.PlaySound(full_file_path, winsound.SND_ASYNC)     
        

    def openSecondWindow(self):
        if ("Main" in dir(self)):
            if (self.Main.isVisible()):
                pass
            else:
                self.Main.setVisible(True)
        else:
            # opening the second window
            self.Main = Main() # same as the second window class name
            try:
                self.setMatchResult()
                self.weightTxt.setPlainText(self.matchData[str(self.gameNum)]["weight"])
            except:
                pass
        
        
        self.Main.roundNumLbl.setText(str(self.roundNum))
        
        
    def setPortName(self):
        self._getserial_ports()
        self.portName = (self.portsCbx.currentText()).lower()
        if (self.portName != "__"):
            self.serial = QtSerialPort.QSerialPort(self.portName,
                                                    baudRate=QtSerialPort.QSerialPort.Baud115200, 
                                                    readyRead=self.receive)
            self.connectBtn.setEnabled(True)
        else: 
            self.connectBtn.setEnabled(False)
        print(self.portName)

    @QtCore.pyqtSlot()
    def receive(self):
        # "x12x2x05c:cf:7f:c9:1b:"
        while self.serial.canReadLine():            
            self.text = self.serial.readLine().data().decode()
            if (self.text[0]=="x"):
                self.textSplit = self.text.split('x')[1:]
                if (self.controlDialog and self.controlDialog.isVisible()):
                    self.controlDialog.getControllerVal(self.textSplit)
                else:
                    self.ardo()
            self.consoleTxt.setPlainText("received message:\n"+self.text+"\n")
            
            

    @QtCore.pyqtSlot()
    def on_toggled(self):
        if not self.serial.isOpen():
            try:
                self.serial.open(QtCore.QIODevice.ReadWrite)
                if(self.serial.isOpen()):
                    print("connected")
                    self.sendBtn.setEnabled(True)
                    self.portsCbx.setEnabled(False)
                    self.connectBtn.setText("قطع اتصال")
                    self.connectionLbl.setProperty("checked","true")
                    self.connectionLbl.style().unpolish(self.connectionLbl)
                    self.connectionLbl.style().polish(self.connectionLbl)
                    self.connectionLbl.update()
                else:
                    print("not connected")
                    QMessageBox.about(self, "خطا", "امکان برقراری ارتباط وجود ندارد!")
                    self.connectionLbl.setProperty("checked","false")
                    self.connectionLbl.style().unpolish(self.connectionLbl)
                    self.connectionLbl.update()
            except:
                QMessageBox.about(self, "خطا", "امکان برقراری ارتباط وجود ندارد!")
                pass
                
        else:
            self.connectionLbl.setProperty("checked","false")
            self.connectionLbl.style().unpolish(self.connectionLbl)
            self.connectionLbl.update()
            self.serial.close()
            self.portsCbx.setEnabled(True)
            self.connectBtn.setText("اتصال")
            self.sendBtn.setEnabled(False)
            self._getserial_ports()
                    
    @QtCore.pyqtSlot()
    def sendToArdo(self, packet=""):
        if (packet == ""):
            packet = self.pointPacket()
            if self.serial.isOpen():
                self.consoleTxt.clear()
                self.serial.write(packet.encode())
                self.consoleTxt.setPlainText("sent message:\n"+packet+"\n")
        else:
            if self.serial.isOpen():
                self.consoleTxt.clear()
                self.serial.write(packet.encode())
                self.consoleTxt.setPlainText("sent message:\n"+packet+"\n")
    
    @QtCore.pyqtSlot()
    def pointPacket(self):
        packet = "9"
        for idx in range(5):
            # tmp = "{}{}{:02d}{:02d}".format("9",str(self.refColor[str(self.roundNum)][idx]),
            #                                 int(self.redPlayerRefsPoint[str(self.roundNum)][idx]),
            #                                 int(self.bluePlayerRefsPoint[str(self.roundNum)][idx]) )
            tmp = "{}".format(str(self.refColor[str(self.roundNum)][idx]))
            packet += tmp
        packet += "\n"
        # if self.serial.isOpen():
        #     self.consoleTxt.clear()
        #     self.serial.write(packet.encode())
        #     self.consoleTxt.setPlainText("sent message:\n"+packet+"\n")
        return packet

    @QtCore.pyqtSlot()
    def ardo(self):
        try:
            refIdx = int(self.refdic[self.textSplit[2][1:16]])-1
            refRedLcd = redLcdList[refIdx]
            refBlueLcd = blueLcdList[refIdx]
            redPoint = int(self.textSplit[0])
            bluePoint = int(self.textSplit[1])
            myObjRed = eval("self."+refRedLcd)
            myObjBlue = eval("self."+refBlueLcd)
            myObjRed.display(str(redPoint))
            myObjBlue.display(str(bluePoint))
        except Exception as ex:
            print(ex)
            QMessageBox.about(self, "خطا", "خطا در دریافت اطلاعات از دسته رخ داده است. تنظیمات دسته ها را بررسی کنید.")

        


        
    def setLcdColor(self):
        self.setStackedWidgetColor(self.lcdNumber, self.lcdNumber_2, self.stackedWidget, refIdx=0)
        self.setStackedWidgetColor(self.lcdNumber_3, self.lcdNumber_4, self.stackedWidget_3, refIdx=1)
        self.setStackedWidgetColor(self.lcdNumber_7, self.lcdNumber_8, self.stackedWidget_4, refIdx=2)
        self.setStackedWidgetColor(self.lcdNumber_9, self.lcdNumber_10, self.stackedWidget_5, refIdx=3)
        self.setStackedWidgetColor(self.lcdNumber_13, self.lcdNumber_14, self.stackedWidget_6, refIdx=4)



    def setStackedWidgetColor(self, lcdRed, lcdBlue, myStackedWidget, refIdx):
        if (int(lcdRed.value()) > int(lcdBlue.value())):
            myStackedWidget.setCurrentIndex(colorDict["red"])
            self.refColor[str(self.roundNum)][int(refIdx)] = colorDict["red"]
        elif (int(lcdRed.value()) < int(lcdBlue.value())):
            myStackedWidget.setCurrentIndex(colorDict["blue"])
            self.refColor[str(self.roundNum)][int(refIdx)] = colorDict["blue"]
        else:
            self.refColor[str(self.roundNum)][int(refIdx)] = colorDict["white"]
            myStackedWidget.setCurrentIndex(colorDict["white"])

    def uploadData(self):
        if ("Main" not in dir(self)):
            QMessageBox.about(self, "اخطار", "لطفا اسکوربورد را باز کنید.")
        else:
            self.calcPoints()
            self.storeTableData()
            try:
                self.myDataTable[int(self.gameNum)-1]["redRefPoint"] = copy.deepcopy(self.redPlayerRefsPoint)
                self.myDataTable[int(self.gameNum)-1]["redTablePoint"] = copy.deepcopy(self.redPlayerTablePoint)
                self.myDataTable[int(self.gameNum)-1]["blueRefPoint"] = copy.deepcopy(self.bluePlayerRefsPoint)
                self.myDataTable[int(self.gameNum)-1]["blueTablePoint"] = copy.deepcopy(self.bluePlayerTablePoint)
            except Exception as ex:
                print("line {}: {}".format(str(inspect.currentframe().f_lineno),ex))
                pass
            self.sendToArdo()
            if (self.redPoints > self.bluePoints):
                myObj = eval("self.Main.stackedWidgetRed"+str(self.roundNum))
                myObj.setCurrentIndex(1)
                myObj = eval("self.Main.stackedWidgetBlue"+str(self.roundNum))
                myObj.setCurrentIndex(0)
                self.redWinRand[int(self.roundNum-1)] = 1
                self.blueWinRand[int(self.roundNum-1)] = 0

            elif (self.redPoints < self.bluePoints):
                myObj = eval("self.Main.stackedWidgetBlue"+str(self.roundNum))
                myObj.setCurrentIndex(1)
                myObj = eval("self.Main.stackedWidgetRed"+str(self.roundNum))
                myObj.setCurrentIndex(0)
                self.redWinRand[int(self.roundNum-1)] = 0
                self.blueWinRand[int(self.roundNum-1)] = 1
            else:
                myObj = eval("self.Main.stackedWidgetBlue"+str(self.roundNum))
                myObj.setCurrentIndex(0)
                myObj = eval("self.Main.stackedWidgetRed"+str(self.roundNum))
                myObj.setCurrentIndex(0)
            self.bluePoints = 0
            self.redPoints = 0
            
            if (sum(self.redWinRand)==2):
                self.Main.winnerStackedWidget.setCurrentIndex(0)
                self.winnerStackedWidget.setCurrentIndex(0)
            elif (sum(self.blueWinRand)==2):
                self.Main.winnerStackedWidget.setCurrentIndex(2)
                self.winnerStackedWidget.setCurrentIndex(2)
            else:
                self.Main.winnerStackedWidget.setCurrentIndex(1)
                self.winnerStackedWidget.setCurrentIndex(1)

    def resetRefLcd(self):
        children = self.findChildren(QLCDNumber)
        for child in children:
            child.display("0")

    def storeTableData(self):
        for idx in range(5):
            for round in range(1,4):
                myObj = eval("self.red"+str(round)+"_"+str(idx+1) + "spinBox")
                self.redPlayerTablePoint[str(round)][idx] =  int(myObj.value())

                myObj = eval("self.blue"+str(round)+"_"+str(idx+1) + "spinBox")
                self.bluePlayerTablePoint[str(round)][idx] =  int(myObj.value())

    


    def calcPoints(self):
        self.bluePoints = 0
        self.redPoints = 0
        for idx, item in enumerate(redLcdList):
            myObjRed = eval("self."+item)
            myObjBlue = eval("self."+blueLcdList[idx])
            self.redPlayerRefsPoint[str(self.roundNum)][idx] = int(myObjRed.value())
            self.bluePlayerRefsPoint[str(self.roundNum)][idx] = int(myObjBlue.value())
            
            if (int(myObjRed.value()) > int(myObjBlue.value())):
                self.redPoints += 1
            elif (int(myObjRed.value()) < int(myObjBlue.value())):
                self.bluePoints += 1
            else:
                self.redPoints += 1
                self.bluePoints += 1
        if ("Main" in dir(self)):
            myProgressVal = int((float(self.redPoints)/( float(self.redPoints) +  float(self.bluePoints)))*100)
            self.Main.progressBar.setValue(myProgressVal)



    def loadCsvFile(self):
        path = self.getfile()
        if (path!=""):
            self.redPlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
            self.bluePlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}

            self.redPlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
            self.bluePlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
            self.loadingCsv = True
            self.openExcelPandas(path)
            self.gameNumCbx.clear()            
            self.gameNumCbx.addItems(map(str, self.matchData.keys()))
            numOfMatches = len(self.matchData)
            self.myDataTable = list()
            for idx in range(numOfMatches):
                tmp = {}
                tmp["redRefPoint"] = copy.deepcopy(self.redPlayerRefsPoint)
                tmp["redTablePoint"] = copy.deepcopy(self.redPlayerTablePoint)
                tmp["blueRefPoint"] = copy.deepcopy(self.bluePlayerRefsPoint)
                tmp["blueTablePoint"] = copy.deepcopy(self.bluePlayerTablePoint)
                self.myDataTable.append(copy.deepcopy(tmp))
            self.gameNum = self.gameNumCbx.currentText()
            self.isMatchData = True
            self.resetRefLcd()
            self.setMatchResult()
        else:
            pass

    def openExcelPandas(self, path):
        df = pd.read_excel(path)
        self.matchData = {}
        
        for idx, item in df.iterrows():
            tmp = {}
            tmp["red"] = copy.deepcopy(item.get("قرمز",""))
            tmp["blue"] = copy.deepcopy(item.get("آبی", ""))
            tmp["weight"] = copy.deepcopy(str(item.get("وزن", "")))
            self.matchData[str(item.get("شماره مسابقه", ""))] = copy.deepcopy(tmp)
        
        if ("" in self.matchData.keys()):
            QMessageBox.about(self, "خطا", "لطفا فایل با قالب مناسب بارگزاری کنید!")
        
        

    def open_workbook(self, path):
        workbook = openpyxl.load_workbook(filename=path)
        print(f"Worksheet names: {workbook.sheetnames}")
        sheet = workbook.active
        print(f"The title of the Worksheet is: {sheet.title}")
        print(f"Cells that contain data: {sheet.calculate_dimension()}")
        for row in sheet.rows:
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Skip this cell
                    continue
                print(f"{cell.column_letter}{cell.row} = {cell.value}")

    def getfile(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        fname = QFileDialog.getOpenFileName(self, 'Open file', 
         dir_path,"Excel files (*.csv *.XLSX)")
        if (fname[0]!= ""):
            fname = fname[0].replace("/","\\")
            return(fname)
        else:
            return("")

    def _getserial_ports(self):
        ports = QtSerialPort.QSerialPortInfo.availablePorts()
        for port in ports :
            if (port.portName() not in self.portsList):
                self.portsCbx.addItem(str(port.portName()))
                self.portsList.append(copy.copy(port.portName()))
    
    def up1r(self):
        v = self.lcdNumber.value()
        self.lcdNumber.display(v+1)  
    def up2r(self):
        v = self.lcdNumber_3.value()
        self.lcdNumber_3.display(v+1) 
    def up3r(self):
        v = self.lcdNumber_7.value()
        self.lcdNumber_7.display(v+1) 
    def up4r(self): 
        v = self.lcdNumber_9.value()
        self.lcdNumber_9.display(v+1)
    def up5r(self):       
        v = self.lcdNumber_13.value()
        self.lcdNumber_13.display(v+1)

    def up1b(self): 
        v = self.lcdNumber_2.value()
        self.lcdNumber_2.display(v+1)
    def up2b(self): 
        v = self.lcdNumber_4.value()
        self.lcdNumber_4.display(v+1)
    def up3b(self):
        v = self.lcdNumber_8.value()
        self.lcdNumber_8.display(v+1) 
    def up4b(self): 
        v = self.lcdNumber_10.value()
        self.lcdNumber_10.display(v+1)
    def up5b(self): 
        v = self.lcdNumber_14.value()
        self.lcdNumber_14.display(v+1)

    def dn1r(self):
        v = self.lcdNumber.value()
        if (v > 0):
            self.lcdNumber.display(v-1)
    def dn2r(self):
        v = self.lcdNumber_3.value()
        if (v > 0):
            self.lcdNumber_3.display(v-1)
    def dn3r(self):
        v = self.lcdNumber_7.value()
        if (v > 0):
            self.lcdNumber_7.display(v-1)
    def dn4r(self):
        v = self.lcdNumber_9.value()
        if (v > 0):
            self.lcdNumber_9.display(v-1)
    def dn5r(self):
        v = self.lcdNumber_13.value()
        if (v > 0):
            self.lcdNumber_13.display(v-1)
    
    def dn1b(self):
        v = self.lcdNumber_2.value()
        if (v > 0):
            self.lcdNumber_2.display(v-1)
    def dn2b(self):
        v = self.lcdNumber_4.value()
        if (v > 0):
            self.lcdNumber_4.display(v-1)
    def dn3b(self):
        v = self.lcdNumber_8.value()
        if (v > 0):
            self.lcdNumber_8.display(v-1)
    def dn4b(self):
        v = self.lcdNumber_10.value()
        if (v > 0):
            self.lcdNumber_10.display(v-1)
    def dn5b(self):
        v = self.lcdNumber_14.value()
        if (v > 0):
            self.lcdNumber_14.display(v-1)


    def closeEvent(self, event):
        close = QtWidgets.QMessageBox.question(self,
                                         "خروج",
                                         "آیا برای خروج از برنامه اطمینان دارید؟",
                                         QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if close == QtWidgets.QMessageBox.Yes:
            if ("Main" in dir(self)):
                self.Main.close()
            if(self.serial.isOpen()):
                self.serial.close()
            event.accept()
        else:
            event.ignore()

    def resetGame(self):
        if (self.loadingCsv == False):
            changeGame = QtWidgets.QMessageBox.question(self,
                                            " شروع مجدد",
                                            "آیا برای شروع مجدد اطمینان دارید؟",
                                            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        else:
            changeGame = QtWidgets.QMessageBox.Yes
            self.loadingCsv = False
        
        
        if changeGame == QtWidgets.QMessageBox.Yes:
            if (self.gameNum == ""):
                self.redWinRand = [0, 0 , 0]
                self.blueWinRand = [0, 0 , 0]
                self.redPlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                self.bluePlayerRefsPoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}

                self.redPlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                self.bluePlayerTablePoint = {"1":[0,0,0,0,0], "2":[0,0,0,0,0], "3":[0,0,0,0,0]}
                try:
                    self.setRefLcd()
                    self.setTableVal()
                except Exception as ex:
                    print("line {}: {}".format(str(inspect.currentframe().f_lineno),ex))
                    pass
                if ("Main" in dir(self)):
                    self.Main.gameNumTxt.setHtml("<p align='center'>" + str(self.gameNum) + "</p>")
                    self.setPlayerName()
                    for tmpRoundNum in range(1,4):
                        myObj = eval("self.Main.stackedWidgetBlue"+str(tmpRoundNum))
                        myObj.setCurrentIndex(0)
                        myObj = eval("self.Main.stackedWidgetRed"+str(tmpRoundNum))
                        myObj.setCurrentIndex(0)
                    self.Main.winnerStackedWidget.setCurrentIndex(1)
                    self.winnerStackedWidget.setCurrentIndex(1)




def splashLoadingScreen():
    splash_object = QSplashScreen(QtGui.QPixmap(":/png_files/Wushu.png"))
    opaqueness = 0.05
    step = 0.05
    sleepTime = 0.05
    splash_object.setWindowOpacity(opaqueness)
    splash_object.show()
    loadingTime = 4
    for _ in range(loadingTime):
        while ((opaqueness < 1) & (opaqueness > 0)):
            splash_object.setWindowOpacity(opaqueness)
            time.sleep(sleepTime) # Gradually appears
            opaqueness+=step
        step = (-1) * step
        opaqueness+=step
    time.sleep(1) # hold image on screen for a while
    splash_object.close() # close the splash screen


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('Fusion')
    # splashLoadingScreen()
    Ui =  MainWindow()
    Ui.show()
    sys.exit(app.exec_())